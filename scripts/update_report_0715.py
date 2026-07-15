"""
7.15 航管会议纪要 → 墨西哥航线运价报表 增量更新
直接修改已生成的报表，覆盖/补充数据，不新建多余 Sheet。
"""
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

SRC = r"C:\Users\12724\Desktop\墨西哥航线运价报表.xlsx"
OUT = r"C:\Users\12724\Desktop\墨西哥航线运价报表_v2.xlsx"

# ── 样式 ──────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
DATA_FONT = Font(name="微软雅黑", size=9)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
UPDATE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 黄色标记会议更新
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BEST_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

wb = load_workbook(SRC)

# ════════════════════════════════════════════════════════
# 会议新增/覆盖数据（人工梳理后的结构化记录）
# ════════════════════════════════════════════════════════

MEETING_DATA = [
    # ── 大连 (新起运港) ──
    # POL, POD, 船公司, 20GP, 40GP, 免柜期, 有效期, 航程, 备注
    ["大连", "Manzanillo", "COSCO", None, None, "", "", "", "20GP<15tons；集团约再看；NOR 放的慢"],
    ["大连", "Manzanillo", "MSC", None, None, "", "", "", "建议走集团Basket"],
    ["大连", "Manzanillo", "HMM", None, None, "", "", "", "降价幅度较慢"],

    # ── 上海 ──
    ["Shanghai", "Manzanillo", "COSCO", None, None, "", "", "", "成本≈4900，预计还会跌 400-500"],
    ["Shanghai", "Manzanillo", "HPL", None, 4100, "", "尽快订", "", "葫芦娃 MX 4100/40HQ，舱位不多"],
    ["Shanghai", "Manzanillo", None, None, None, "", "", "", "⚠ 台风影响，船期混乱，多船跳挂，提前告知代理"],

    # ── 宁波 ──
    ["Ningbo", "Manzanillo", "MSC", None, None, "", "", "", "NOR 缺箱；7.15/26 空班"],
    ["Ningbo", "Manzanillo", "COSCO", None, None, "", "", "", "WSA 航线 open"],
    ["Ningbo", "Manzanillo", "PIL", None, None, "", "", "", "箱子紧张，有打到单子尽快"],
    ["Ningbo", "Manzanillo", "EMC", None, None, "", "", "", "看群信息刷箱，箱子紧张"],
    ["Ningbo", "Manzanillo", "ONE", None, None, "", "", "", "看现舱，群里更新"],
    ["Ningbo", "Manzanillo", "TSL", None, None, "", "", "", "MX 可申请目的港免押箱费"],

    # ── 天津 ──
    ["Tianjin", "Manzanillo", "MSK", None, None, "", "", "", "依旧推荐 MSK，价格继续下调；零散单首选"],
    ["Tianjin", "Manzanillo", "MSC", 4000, 4100, "", "", "", "NAC 成本 4000/4100；每周五直航；小重柜南美西可订"],
    ["Tianjin", "Manzanillo", "CMA", None, None, "", "", "", "看群更新价格；7.21 加勒比 +50/100；7.25 后附加费 40→50 USD"],
    ["Tianjin", "Manzanillo", "MSC", None, None, "", "", "", "27 号放舱政策已出；加勒比总体比 CMA 低"],
    ["Tianjin", "Manzanillo", "COSCO", None, None, "", "", "", "7.22/26 未绑拖车已报备未批复；7.29/30 绑拖车要求可能放宽"],
    ["Tianjin", "Manzanillo", "EMC", None, None, "", "", "", "头程船司变化频繁；加勒比限重 20GP<10T, 40HQ<20T（比之前好）"],
    ["Tianjin", "Manzanillo", "ZIM", None, None, "", "", "", "加勒比 7.21/28 船，价格偏贵"],
    ["Tianjin", "Manzanillo", None, None, None, "", "", "", "MX 要 3 开头海运费就走 NOR；2xxx/3xxx 只有特价舱"],

    # ── Progreso 补充 ──
    ["Tianjin", "Progreso", "MSC", None, None, "", "27 号放舱", "", "Progreso MSC 7 月 NAC 正在申请暂未批复"],
    ["Shanghai", "Progreso", "PIL", None, None, "", "7.30 船", "", "南美东看内训单"],

    # ── Lazaro Cardenas 补充 ──
    ["Tianjin", "Lazaro Cardenas", "MSK", None, None, "", "", "", "天津 MSK 线上拍仓需跟航管报备"],

    # ── 全局备注 ──
    ["(全局)", "(全局)", None, None, None, "", "", "", "大票货找 Lucas 单询"],
    ["(全局)", "(全局)", None, None, None, "", "", "", "海外目的港问题可找 Daisy"],
]


def find_row_range(ws, pol, pod):
    """在明细 sheet 中找到某个 POL→POD 分组的行范围"""
    # ws row 4 = header, row 5+ = data
    target_section = None
    for row in range(5, ws.max_row + 1):
        cell_val = str(ws.cell(row=row, column=1).value or "")
        if cell_val.startswith("▎"):
            target_section = cell_val.replace("▎", "").strip()
            continue
        if target_section == pol:
            pod_cell = str(ws.cell(row=row, column=3).value or "")
            if pod.lower() in pod_cell.lower():
                return row
    return None


def insert_row(ws, row, values, highlight=True):
    """在指定行插入新行"""
    ws.insert_rows(row)
    for j, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=j, value=v if v is not None else "")
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if j in (1, 2, 3):
            cell.alignment = CENTER
        elif j in (4, 5):
            cell.alignment = CENTER
            if isinstance(v, (int, float)):
                cell.number_format = '#,##0'
        else:
            cell.alignment = LEFT_WRAP
        if highlight:
            cell.fill = UPDATE_FILL


def find_last_data_row_in_section(ws, pol):
    """找到某个 POL 分组下最后一条数据行"""
    in_section = False
    last_row = None
    for row in range(5, ws.max_row + 1):
        cell_val = str(ws.cell(row=row, column=1).value or "")
        if cell_val.startswith("▎"):
            if in_section:
                break  # 进入下一个 section
            if cell_val.replace("▎", "").strip() == pol:
                in_section = True
            continue
        if in_section:
            carrier = ws.cell(row=row, column=1).value
            if carrier and str(carrier).strip():
                last_row = row
    return last_row


def add_or_update_in_detail_sheet(ws, pod_name, pol, carrier, gp20, gp40, ft, validity, tt, remark):
    """在明细 sheet 中，在同 POL 分组内找到对应船公司行，存在则追加备注；不存在则插入"""
    header_row = 4
    ncols = 9

    # 先找该 POL section
    target_section_start = None
    next_section_start = None
    for row in range(5, ws.max_row + 2):
        cell_val = str(ws.cell(row=row, column=1).value or "")
        if cell_val.startswith("▎"):
            section_pol = cell_val.replace("▎", "").strip()
            if section_pol == pol:
                target_section_start = row
            elif target_section_start and section_pol != pol:
                next_section_start = row
                break

    if target_section_start is None:
        return  # POL 不存在

    # 搜索范围
    search_end = next_section_start if next_section_start else ws.max_row + 1

    # 在 section 内找对应船公司
    for row in range(target_section_start + 1, search_end):
        existing_carrier = str(ws.cell(row=row, column=1).value or "").strip()
        if existing_carrier.upper() == carrier.upper():
            # 已存在 → 追加备注
            if remark:
                old_remark = str(ws.cell(row=row, column=9).value or "")
                new_remark = (old_remark + "\n" + remark).strip()
                ws.cell(row=row, column=9, value=new_remark)
                ws.cell(row=row, column=9).fill = UPDATE_FILL
            # 如果会议给了新价格，覆盖空值
            if gp20 is not None and ws.cell(row=row, column=4).value in (None, "", "—"):
                ws.cell(row=row, column=4, value=gp20)
                ws.cell(row=row, column=4).fill = UPDATE_FILL
            if gp40 is not None and ws.cell(row=row, column=5).value in (None, "", "—"):
                ws.cell(row=row, column=5, value=gp40)
                ws.cell(row=row, column=5).fill = UPDATE_FILL
            return

    # 不存在 → 在 section 末尾插入
    insert_pos = find_last_data_row_in_section(ws, pol)
    if insert_pos is None:
        insert_pos = target_section_start + 1

    values = [carrier, pol, pod_name, gp20 if gp20 else "—", gp40 if gp40 else "—",
              ft, validity, tt, remark]
    insert_row(ws, insert_pos + 1, values, highlight=True)


def update_detail_sheet(ws_name, pod_name):
    """更新一个明细 sheet"""
    if ws_name not in wb.sheetnames:
        return
    ws = wb[ws_name]

    for rec in MEETING_DATA:
        r_pol, r_pod, r_carrier, r_20, r_40, r_ft, r_valid, r_tt, r_remark = rec
        if r_pod != pod_name:
            continue
        if r_carrier is None:
            # 全局备注 → 插在 POL 分组末尾
            if r_pol == "(全局)":
                continue
            insert_pos = find_last_data_row_in_section(ws, r_pol)
            if insert_pos:
                values = ["📋", r_pol, pod_name, "—", "—", "", "", "", r_remark]
                insert_row(ws, insert_pos + 1, values, highlight=True)
            continue
        add_or_update_in_detail_sheet(ws, pod_name, r_pol, r_carrier,
                                      r_20, r_40, r_ft, r_valid, r_tt, r_remark)


# ── 1. 更新三个明细 Sheet ──
for sheet_name, pod in [("Manzanillo", "Manzanillo"),
                         ("Lazaro Cardenas", "Lazaro Cardenas"),
                         ("Progreso", "Progreso")]:
    update_detail_sheet(sheet_name, pod)

# ── 2. 新增大连起运港（三个港都加） ──
DALIAN_DATA = [
    # POD, carrier, 20gp, 40gp, free_time, validity, tt, remark
    ("Manzanillo", "COSCO", None, None, "", "", "", "20GP<15tons；集团约再看；NOR 放的慢"),
    ("Manzanillo", "MSC", None, None, "", "", "", "建议走集团 Basket"),
    ("Manzanillo", "HMM", None, None, "", "", "", "降价幅度较慢"),
    ("Lazaro Cardenas", "COSCO", None, None, "", "", "", "20GP<15tons；集团约再看"),
    ("Lazaro Cardenas", "MSC", None, None, "", "", "", "建议走集团 Basket"),
    ("Progreso", "COSCO", None, None, "", "", "", "20GP<15tons；集团约再看"),
    ("Progreso", "MSC", None, None, "", "", "", "建议走集团 Basket"),
]

for pod_name in ["Manzanillo", "Lazaro Cardenas", "Progreso"]:
    if pod_name not in wb.sheetnames:
        continue
    ws = wb[pod_name]
    ncols = 9

    # 在最后一个 POL 分组后插入大连 section
    # 找最后一行
    last_row = ws.max_row
    insert_at = last_row + 2

    # 插入分隔行
    ws.merge_cells(start_row=insert_at, start_column=1, end_row=insert_at, end_column=ncols)
    cell = ws.cell(row=insert_at, column=1, value="▎大连（7.15 新增）")
    cell.font = SECTION_FONT
    for c in range(1, ncols + 1):
        ws.cell(row=insert_at, column=c).fill = UPDATE_FILL
        ws.cell(row=insert_at, column=c).border = THIN_BORDER

    r = insert_at + 1
    for d in DALIAN_DATA:
        if d[0] != pod_name:
            continue
        values = [d[1], "大连", d[0], "待确认" if d[2] is None else d[2],
                  "待确认" if d[3] is None else d[3], d[4], d[5], d[6], d[7]]
        for j, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.fill = UPDATE_FILL
            cell.alignment = CENTER if j <= 3 else LEFT_WRAP
        r += 1

# ── 3. 更新汇总 Sheet ──
ws_summary = wb["运价汇总"]

# 在汇总中新增大连条目（仅 Manzanillo 有 MSC NAC 价）
last_summary_row = ws_summary.max_row
for rec in MEETING_DATA:
    r_pol, r_pod, r_carrier, r_20, r_40, _, _, _, r_remark = rec
    if r_pol != "大连" or r_carrier != "MSC" or r_pod != "Manzanillo":
        continue
    # 在 Manzanillo 分组末尾插入
    values = [r_pod, r_pol, r_carrier, r_20 if r_20 else "—", r_40 if r_40 else "—", "", "", "", r_remark]
    # 找 Manzanillo 最后一条
    insert_at = None
    for row in range(5, ws_summary.max_row + 1):
        pod_val = str(ws_summary.cell(row=row, column=1).value or "")
        if pod_val == "Manzanillo":
            insert_at = row
    if insert_at:
        insert_at += 1
    else:
        insert_at = last_summary_row + 1

    ws_summary.insert_rows(insert_at)
    for j, v in enumerate(values, 1):
        cell = ws_summary.cell(row=insert_at, column=j, value=v)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.fill = UPDATE_FILL
        cell.alignment = CENTER if j <= 3 else LEFT_WRAP

# ── 4. 更新说明 Sheet ──
ws_note = wb.create_sheet(title="7.15更新说明")
ws_note.merge_cells("A1:C1")
ws_note.cell(row=1, column=1, value="7.15 航管会议 — 更新说明").font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
ws_note.merge_cells("A2:C2")
ws_note.cell(row=2, column=1, value="黄色高亮行 = 会议新增/补充数据").font = Font(name="微软雅黑", size=9, color="808080")

notes = [
    ["大连", "新增起运港", "COSCO / MSC / HMM 三条航线，具体运价待确认"],
    ["上海 COSCO", "备注补充", "成本≈4900，预计继续跌 400-500"],
    ["上海 HPL", "新增报价", "葫芦娃 MX 4100/40HQ，舱位不多"],
    ["上海全局", "备注补充", "台风影响，船期混乱，多船跳挂"],
    ["宁波 TSL", "备注补充", "MX 可申请目的港免押箱费"],
    ["宁波 EMC/PIL", "备注补充", "箱子紧张"],
    ["天津 MSC", "新增 NAC 价", "4000/4100（南美西 NAC 成本价）"],
    ["天津 MSK", "备注补充", "依旧推荐，价格继续下调；线上拍仓需报备"],
    ["天津 CMA", "备注补充", "7.21 加勒比 +50/100；7.25 后附加费 40→50 USD"],
    ["天津 EMC", "备注补充", "加勒比限重 20GP<10T, 40HQ<20T"],
    ["天津全局", "备注补充", "要 3 开头就走 NOR"],
    ["全局", "备注补充", "大票货找 Lucas 单询；海外找 Daisy"],
]
for i, (a, b, c) in enumerate(notes, 4):
    ws_note.cell(row=i, column=1, value=a).font = DATA_FONT
    ws_note.cell(row=i, column=2, value=b).font = DATA_FONT
    ws_note.cell(row=i, column=3, value=c).font = DATA_FONT
    for j in range(1, 4):
        ws_note.cell(row=i, column=j).border = THIN_BORDER
ws_note.column_dimensions["A"].width = 18
ws_note.column_dimensions["B"].width = 16
ws_note.column_dimensions["C"].width = 55

# ── 保存 ──
wb.save(OUT)
print("✅ 报表已更新（覆盖/补充模式）")
print(f"  保存至: {OUT}")
print("  - 三个明细 Sheet：补充会议数据 + 新增大连起运港")
print("  - 运价汇总：新增大连 MSC Manzanillo 条目")
print("  - 7.15更新说明：新增 Sheet 记录本次变更")
print("  - 黄色高亮 = 本次会议新增/修改的内容")
