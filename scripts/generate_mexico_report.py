"""
墨西哥航线运价报表生成脚本
从 Quotationss.xlsx 提取墨西哥三港（Manzanillo / Lazaro Cardenas / Progreso）运价，
输出格式化报表。
"""
import re
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from copy import copy

# ── 配置 ──────────────────────────────────────────────
SRC = r"C:\Users\12724\Desktop\Quotationss.xlsx"
OUT = r"C:\Users\12724\Desktop\墨西哥航线运价报表.xlsx"

CARRIER_MAP = {  # 代码 → 船公司名
    "001": "CMA", "002": "COSCO", "003": "CSSC", "004": "EMC",
    "005": "HMM", "006": "HPL", "007": "MSC", "008": "MSK",
    "009": "ONE", "010": "OOCL", "011": "PIL", "012": "WHL",
    "013": "YML", "014": "ZIM", "015": "BAL", "016": "ESL",
    "017": "SNL", "018": "Sealead", "019": "TSL", "020": "SINOTRANS",
    "021": "RCL", "7": "七家联盟",
    "KMTC/RCL": "KMTC/RCL",
}
# 船公司短代码匹配: "005(MM)" → "005"
RE_CARRIER_CODE = re.compile(r"^(\d+|KMTC/RCL)")

# ── 样式 ──────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=9)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="微软雅黑", size=9, color="808080")
BEST_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 最优价高亮
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")

# ── 数据提取 ──────────────────────────────────────────
def parse_carrier(raw):
    """从 '005(MM)' 或 '001' 中提取纯数字代码，返回 (code, display_name)"""
    if pd.isna(raw) or str(raw).strip() == "":
        return (None, None)
    s = str(raw).strip()
    m = RE_CARRIER_CODE.match(s)
    if m:
        code = m.group(1)
        name = CARRIER_MAP.get(code, code)
        return (code, name)
    return (None, None)


def clean_port(raw):
    """清理港口名: 'Manzanillo，MX' → 'Manzanillo'"""
    if pd.isna(raw) or str(raw).strip() == "":
        return ""
    s = str(raw).strip()
    s = s.replace("\xa0", " ").replace("，", ",")
    s = re.sub(r",\s*MX", "", s, flags=re.IGNORECASE)
    # 统一中文逗号
    return s.strip()


def is_valid_rate(val):
    """判断是否为有效运价数字"""
    if pd.isna(val):
        return False
    s = str(val).strip().lower()
    if s in ("", "/", "rate pending", "-", "n/a"):
        return False
    try:
        float(str(val).replace(",", "").replace("$", ""))
        return True
    except ValueError:
        return False


def parse_rate(val):
    """解析运价为浮点数"""
    if not is_valid_rate(val):
        return None
    return float(str(val).replace(",", "").replace("$", ""))


def normalize_sheet(df):
    """
    把原始 sheet 的杂牌格式标准化为统一 DataFrame。
    原始 header 在 Row 1（第 2 行），数据从 Row 2 开始。
    """
    # 找到 header 行（含 CARRIER）
    header_row = None
    for i in range(min(5, len(df))):
        row_str = " ".join([str(v) for v in df.iloc[i].values if pd.notna(v)])
        if "CARRIER" in row_str.upper():
            header_row = i
            break
    if header_row is None:
        return pd.DataFrame()

    # 用 header 行做列名
    cols = []
    for v in df.iloc[header_row].values:
        c = str(v).strip() if pd.notna(v) else ""
        cols.append(c)
    df = df.iloc[header_row + 1:].copy()
    df.columns = [f"COL_{j}" for j in range(len(cols))]
    # 映射到标准列
    col_map = {}
    for j, c in enumerate(cols):
        cu = c.upper()
        if "CARRIER" in cu:
            col_map["CARRIER"] = f"COL_{j}"
        elif "POL" in cu:
            col_map["POL"] = f"COL_{j}"
        elif "POD" in cu:
            col_map["POD"] = f"COL_{j}"
        elif "20GP" in cu or "20'" in cu:
            col_map["20GP"] = f"COL_{j}"
        elif "40GP" in cu or "40HQ" in cu or "40'" in cu or "40GP/HQ" in cu:
            col_map["40GP"] = f"COL_{j}"
        elif "FREE" in cu:
            col_map["FREE_TIME"] = f"COL_{j}"
        elif "VALID" in cu:
            col_map["VALIDITY"] = f"COL_{j}"
        elif col_map.get("T_T") is None and ("T/T" in c or "T_T" in c or c.strip() == "T/T"):
            col_map["T_T"] = f"COL_{j}"
        elif "REMARK" in cu:
            col_map["REMARK"] = f"COL_{j}"

    records = []
    last_pol = ""
    last_pod = ""
    for _, row in df.iterrows():
        carrier_raw = row.get(col_map.get("CARRIER", ""))
        pol_raw = row.get(col_map.get("POL", ""))
        pod_raw = row.get(col_map.get("POD", ""))
        gp20_raw = row.get(col_map.get("20GP", ""))
        gp40_raw = row.get(col_map.get("40GP", ""))
        ft_raw = row.get(col_map.get("FREE_TIME", ""))
        val_raw = row.get(col_map.get("VALIDITY", ""))
        tt_raw = row.get(col_map.get("T_T", ""))
        remark_raw = row.get(col_map.get("REMARK", ""))

        # 继承上一行的 POL
        pol = clean_port(pol_raw)
        pod = clean_port(pod_raw)
        if pol:
            last_pol = pol
        if pod:
            last_pod = pod

        code, cname = parse_carrier(carrier_raw)
        # 如果该行没有 carrier 也没有运价，跳过
        gp20 = parse_rate(gp20_raw)
        gp40 = parse_rate(gp40_raw)
        if code is None and gp20 is None and gp40 is None:
            continue

        ft = str(ft_raw).strip() if pd.notna(ft_raw) else ""
        val = str(val_raw).strip() if pd.notna(val_raw) else ""
        tt = str(tt_raw).strip() if pd.notna(tt_raw) else ""
        remark = str(remark_raw).strip() if pd.notna(remark_raw) else ""

        # 去掉时间戳格式
        if " 00:00:00" in val:
            val = val.replace(" 00:00:00", "")

        records.append({
            "carrier_code": code or "",
            "carrier_name": cname or "",
            "pol": last_pol if last_pol else pol,
            "pod": last_pod if last_pod else pod,
            "20gp": gp20,
            "40gp": gp40,
            "free_time": ft,
            "validity": val,
            "tt": tt,
            "remark": remark,
        })

    return pd.DataFrame(records)


def load_all_data():
    """加载所有相关 sheet 并合并"""
    xls = pd.ExcelFile(SRC)
    target_sheets = ["备份", "Manzanillo", "Lazaro", "Progreso", "0708IVAN"]
    frames = []
    for sheet in target_sheets:
        if sheet not in xls.sheet_names:
            continue
        df = pd.read_excel(xls, sheet_name=sheet, header=None)
        normalized = normalize_sheet(df)
        if not normalized.empty:
            frames.append(normalized)
    if not frames:
        return pd.DataFrame()
    all_data = pd.concat(frames, ignore_index=True)

    # 过滤：只保留墨西哥三港
    mex_ports = ["Manzanillo", "Lazaro", "LZC", "Progreso"]
    def is_mexico(pod):
        p = str(pod).upper().replace(" ", "")
        return any(mp.upper().replace(" ", "") in p for mp in mex_ports)

    all_data = all_data[all_data["pod"].apply(is_mexico)]

    # 标准化 POD 名称
    def norm_pod(pod):
        p = str(pod).strip()
        pu = p.upper()
        if "LZC" in pu or "LAZARO" in pu:
            return "Lazaro Cardenas"
        if "MANZANILLO" in pu:
            return "Manzanillo"
        if "PROGRESO" in pu:
            return "Progreso"
        return p

    all_data["pod"] = all_data["pod"].apply(norm_pod)

    # 去重：同一船公司 + 同一 POL + 同一 POD + 同一运价 → 保留一条
    all_data = all_data.drop_duplicates(
        subset=["carrier_name", "pol", "pod", "20gp", "40gp"],
        keep="first"
    )

    # 按目的港、起运港、20GP 运价排序
    all_data = all_data.sort_values(["pod", "pol", "20gp"], key=lambda x: x.fillna(99999))

    return all_data


# ── Excel 写入 ────────────────────────────────────────
def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_row(ws, row, ncols, is_best=False):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if col in (4, 5):  # 运价列居中
            cell.alignment = CENTER
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
        elif col in (1, 2, 3):  # 船公司/起运港/目的港
            cell.alignment = CENTER
        else:
            cell.alignment = LEFT_WRAP
        if is_best:
            cell.fill = BEST_FILL


def write_detail_sheet(wb, sheet_name, data, pod_filter):
    """写单个目的港的明细 sheet"""
    ws = wb.create_sheet(title=sheet_name)

    df = data[data["pod"] == pod_filter].copy()
    if df.empty:
        ws.cell(row=1, column=1, value=f"{pod_filter} — 暂无数据").font = DATA_FONT
        return

    headers = ["船公司", "起运港 (POL)", "目的港 (POD)", "20GP (USD)", "40GP/HQ (USD)",
               "免柜期", "有效期", "航程 (T/T)", "备注"]
    ncols = len(headers)

    # 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value=f"墨西哥航线 — {pod_filter} 运价明细").font = TITLE_FONT

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1,
            value=f"数据来源: Quotationss.xlsx | 生成日期: 2026-07-15 | 共 {len(df)} 条报价").font = SUBTITLE_FONT

    # Header (row 4)
    for j, h in enumerate(headers, 1):
        ws.cell(row=4, column=j, value=h)
    style_header(ws, 4, ncols)

    # 按起运港分组
    pol_order = ["Shanghai", "Ningbo", "Qingdao", "Tianjin", "Shenzhen",
                 "Shekou", "Yantian", "Xiamen", "Chongqing", "Jakarta"]
    df["pol_sort"] = df["pol"].apply(
        lambda x: pol_order.index(x) if x in pol_order else 99
    )
    df = df.sort_values(["pol_sort", "20gp"], key=lambda x: x.fillna(99999))

    # 找出每个 (POL) 组内的最低价
    best_20 = {}
    best_40 = {}
    for pol in df["pol"].unique():
        subset = df[df["pol"] == pol]
        valid_20 = subset[subset["20gp"].notna() & (subset["20gp"] > 0)]
        valid_40 = subset[subset["40gp"].notna() & (subset["40gp"] > 0)]
        if not valid_20.empty:
            best_20[pol] = valid_20["20gp"].min()
        if not valid_40.empty:
            best_40[pol] = valid_40["40gp"].min()

    current_row = 5
    last_pol = None
    for _, rec in df.iterrows():
        # 新起运港前插入分隔行
        if rec["pol"] != last_pol:
            if last_pol is not None:
                current_row += 1  # 空行
            last_pol = rec["pol"]
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ncols)
            pol_label = f"▎{rec['pol']}"
            ws.cell(row=current_row, column=1, value=pol_label).font = SECTION_FONT
            ws.cell(row=current_row, column=1).fill = SECTION_FILL
            for c in range(1, ncols + 1):
                ws.cell(row=current_row, column=c).fill = SECTION_FILL
                ws.cell(row=current_row, column=c).border = THIN_BORDER
            current_row += 1

        values = [
            rec["carrier_name"] if rec["carrier_name"] else rec["carrier_code"],
            rec["pol"],
            rec["pod"],
            rec["20gp"] if rec["20gp"] else "—",
            rec["40gp"] if rec["40gp"] else "—",
            rec["free_time"],
            rec["validity"],
            rec["tt"],
            rec["remark"],
        ]
        is_best = (
            (rec["20gp"] and rec["20gp"] == best_20.get(rec["pol"])) or
            (rec["40gp"] and rec["40gp"] == best_40.get(rec["pol"]))
        )
        for j, v in enumerate(values, 1):
            ws.cell(row=current_row, column=j, value=v if v is not None else "")
        style_data_row(ws, current_row, ncols, is_best)
        current_row += 1

    # 列宽
    col_widths = [12, 14, 18, 14, 14, 14, 18, 14, 40]
    for j, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}{current_row - 1}"


def write_summary_sheet(wb, data):
    """写汇总 sheet —— 每个 POL→POD 组合的最低报价"""
    ws = wb.active
    ws.title = "运价汇总"

    headers = ["目的港", "起运港", "船公司", "20GP (USD)", "40GP/HQ (USD)",
               "免柜期", "有效期", "航程", "备注"]
    ncols = len(headers)

    # 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value="墨西哥航线运价汇总表").font = TITLE_FONT

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1,
            value=f"数据来源: Quotationss.xlsx | 生成日期: 2026-07-15 | 币种: USD").font = SUBTITLE_FONT

    # Header
    for j, h in enumerate(headers, 1):
        ws.cell(row=4, column=j, value=h)
    style_header(ws, 4, ncols)

    # 对每个 POL→POD 组合选最优价（20GP 最低优先，无 20GP 则按 40GP）
    summary_rows = []
    for pod in ["Manzanillo", "Lazaro Cardenas", "Progreso"]:
        pod_data = data[data["pod"] == pod]
        for pol in pod_data["pol"].unique():
            pol_data = pod_data[pod_data["pol"] == pol]
            # 找 20GP 最低的
            valid_20 = pol_data[pol_data["20gp"].notna() & (pol_data["20gp"] > 0)]
            if not valid_20.empty:
                # 确保 idxmin 不会遇到全 NA
                try:
                    best = valid_20.loc[valid_20["20gp"].idxmin()]
                    summary_rows.append(best)
                    continue
                except ValueError:
                    pass
            # 无有效 20GP，尝试 40GP
            valid_40 = pol_data[pol_data["40gp"].notna() & (pol_data["40gp"] > 0)]
            if not valid_40.empty:
                try:
                    best = valid_40.loc[valid_40["40gp"].idxmin()]
                    summary_rows.append(best)
                    continue
                except ValueError:
                    pass
            # 都没有价格但有数据，取第一条（展示无价格信息）
            if not pol_data.empty:
                summary_rows.append(pol_data.iloc[0])

    # 排序
    pod_order = {"Manzanillo": 0, "Lazaro Cardenas": 1, "Progreso": 2}
    pol_order = {"Shanghai": 0, "Ningbo": 1, "Qingdao": 2, "Tianjin": 3,
                 "Shenzhen": 4, "Shekou": 5, "Yantian": 6,
                 "Xiamen": 7, "Chongqing": 8, "Jakarta": 9}
    summary_rows.sort(key=lambda r: (
        pod_order.get(r["pod"], 99),
        pol_order.get(r["pol"], 99),
    ))

    current_row = 5
    last_pod = None
    for rec in summary_rows:
        if rec["pod"] != last_pod:
            if last_pod is not None:
                current_row += 1
            last_pod = rec["pod"]
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ncols)
            ws.cell(row=current_row, column=1, value=f"▶ {rec['pod']}").font = SECTION_FONT
            for c in range(1, ncols + 1):
                ws.cell(row=current_row, column=c).fill = SECTION_FILL
                ws.cell(row=current_row, column=c).border = THIN_BORDER
            current_row += 1

        values = [
            rec["pod"],
            rec["pol"],
            rec["carrier_name"] if rec["carrier_name"] else rec["carrier_code"],
            rec["20gp"] if rec["20gp"] else "—",
            rec["40gp"] if rec["40gp"] else "—",
            rec["free_time"],
            rec["validity"],
            rec["tt"],
            rec["remark"],
        ]
        for j, v in enumerate(values, 1):
            ws.cell(row=current_row, column=j, value=v if v is not None else "")
        style_data_row(ws, current_row, ncols, is_best=True)
        current_row += 1

    # 列宽
    col_widths = [18, 14, 12, 14, 14, 14, 18, 14, 45]
    for j, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}{current_row - 1}"


def write_carrier_ref(wb):
    """写船公司代码对照表"""
    ws = wb.create_sheet(title="船公司代码表")
    headers = ["代码", "船公司", "备注"]
    ncols = 3

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value="船公司代码对照表").font = TITLE_FONT

    for j, h in enumerate(headers, 1):
        ws.cell(row=3, column=j, value=h)
    style_header(ws, 3, ncols)

    notes = {
        "MSC": "DT 14天可申请21天", "MSK": "7月舱位紧张，需加钱买",
        "PIL": "免柜期 20天", "ESL": "免柜期 18天",
        "EMC": "免柜期 28天，免押金", "WHL": "需另加 AMS USD30 + ISPS USD14",
    }

    row = 4
    for code, name in sorted(CARRIER_MAP.items(), key=lambda x: x[0]):
        if code in ("7", "KMTC/RCL"):
            continue
        ws.cell(row=row, column=1, value=code).font = DATA_FONT
        ws.cell(row=row, column=2, value=name).font = DATA_FONT
        ws.cell(row=row, column=3, value=notes.get(name, "")).font = DATA_FONT
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).alignment = CENTER if c < 3 else LEFT_WRAP
        row += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    ws.freeze_panes = "A4"


# ── Main ──────────────────────────────────────────────
def main():
    print("正在读取数据...")
    data = load_all_data()
    if data.empty:
        print("错误：未提取到任何墨西哥航线数据")
        sys.exit(1)

    # 统计
    for pod in ["Manzanillo", "Lazaro Cardenas", "Progreso"]:
        cnt = len(data[data["pod"] == pod])
        carriers = data[data["pod"] == pod]["carrier_name"].nunique()
        pols = data[data["pod"] == pod]["pol"].nunique()
        print(f"  {pod}: {cnt} 条报价, {carriers} 家船公司, {pols} 个起运港")

    print(f"\n共 {len(data)} 条墨西哥航线报价，正在生成报表...")

    wb = Workbook()

    # 1. 汇总 sheet
    write_summary_sheet(wb, data)

    # 2-4. 三个目的港明细
    for pod in ["Manzanillo", "Lazaro Cardenas", "Progreso"]:
        write_detail_sheet(wb, pod, data, pod)

    # 5. 船公司代码表
    write_carrier_ref(wb)

    wb.save(OUT)
    print(f"\n✅ 报表已生成: {OUT}")
    print(f"   共 5 个 Sheet：运价汇总 + Manzanillo + Lazaro Cardenas + Progreso + 船公司代码表")


if __name__ == "__main__":
    main()
